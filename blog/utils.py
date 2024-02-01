import openpyxl
import requests
from django.conf import settings
from openpyxl.styles import Font


def send_contact_info_to_telegram_chat(data):
    name = data.get("name")
    email_or_username = data.get("email_or_username")
    subject = data.get("subject")
    message = data.get("message")

    text = (
        f"New message: {subject}\n\n"
        f"From: {name}\n"
        f"Email or telegram: {email_or_username}\n"
        f"Message: {message}"
    )

    requests.post(
        f"https://api.telegram.org/bot{settings.TELEGRAM_BOT_TOKEN}/sendMessage",
        data={
            "chat_id": settings.TELEGRAM_CHAT_ID,
            "text": text,
            "disable_web_page_preview": True,
        },
    )


def export_to_excel(data: list, headings: list, filepath: str) -> None:
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.row_dimensions[1].font = Font(bold=True)

    for col_no, heading in enumerate(headings, start=1):
        sheet.cell(row=1, column=col_no).value = heading

    for row_no, row in enumerate(data, start=2):
        for col_no, cell_value in enumerate(row, start=1):
            sheet.cell(row=row_no, column=col_no).value = cell_value

    wb.save(filepath)
